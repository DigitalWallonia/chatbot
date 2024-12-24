import json
import asyncio
import streamlit as st
import pandas as pd
import semantic_kernel as sk
from semantic_kernel.contents.chat_history import ChatHistory
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from langgraph.graph import StateGraph, MessagesState, START, END
from langchain_core.messages import HumanMessage, AIMessage, trim_messages


from utils.import_ttl import import_ttl
from utils.chat_functions import chat_with_index
from utils.taxo_mapping import mapping_taxonomies, chatting_and_mapping_taxonomies
from RAG.agents.supervisor_agent import build_and_test_graph

VARIABLES_DATA = [["Class URI", "http://mapping.D4W.com/entity1alignment", ""],["PREFIX", "mapping", "http://mapping.D4W.com/entity1alignment/"],["PREFIX", "align", "http://knowledgeweb.semanticweb.org/heterogeneity/alignment#"],["PREFIX", "skos", "http://www.w3.org/2004/02/skos/core#"],["rdf:type", "align:Alignment", ""],["", "", ""]]
VARIABLES_DATA_1 = [["Class URI", "http://mapping.D4W.com/entity1", ""],["PREFIX", "align", "http://knowledgeweb.semanticweb.org/heterogeneity/alignment#"],["PREFIX", "skos", "http://www.w3.org/2004/02/skos/core#"],["", "", ""]]
VARIABLES_DATA_2 = [["Class URI", "http://mapping.D4W.com/entity2", ""],["PREFIX", "align", "http://knowledgeweb.semanticweb.org/heterogeneity/alignment#"],["PREFIX", "skos", "http://www.w3.org/2004/02/skos/core#"],["", "", ""]]


def write_excel(df1, df2, df3, file):
    """
    Writes multiple DataFrames to an Excel file with specific sheets and variables data.  
  
    :param df1: The first DataFrame to write.  
    :param df2: The second DataFrame to write.  
    :param df3: The third DataFrame to write.  
    :param file: The file path where the Excel file will be saved.
    """
    with pd.ExcelWriter(file, engine="openpyxl") as writer:
        Workbook = writer.book
        ws1 = Workbook.create_sheet("alignement")
        ws2 = Workbook.create_sheet("entity1")
        ws3 = Workbook.create_sheet("entity2")

        for row1 in VARIABLES_DATA:
            ws1.append(row1)

        for row2, row3 in zip(VARIABLES_DATA_1, VARIABLES_DATA_2):
            ws2.append(row2)
            ws3.append(row3)

        for row1, row2, row3 in zip(dataframe_to_rows(df1, index=False, header=True), dataframe_to_rows(df2, index=False, header=True), dataframe_to_rows(df3, index=False, header=True)):
            ws1.append(row1)
            ws2.append(row2)
            ws3.append(row3)


def set_chatbox_layout(chosen_function):  
    """
    Sets up the chatbox layout in the Streamlit interface.  
  
    :param chosen_function: The chosen function for the chatbox.
    """
    if "upload_bool" not in st.session_state:  
        st.session_state["upload_bool"] = False  
    
    # Title  
    st.title(chosen_function)

    if "history" not in st.session_state:  
        st.session_state["history"] = ChatHistory()  
        message = "Hello there! Importez une taxonomie au format TTL pour commencer cette session."  
        st.session_state["history"].add_assistant_message(message)
        st.session_state.chat_history = []
        st.session_state.chat_history.append(AIMessage(content=message))
    

    # Import section  
    label = "Importer votre taxonomie (au format TTL)"  
    uploaded_file = st.file_uploader(label, type="ttl", accept_multiple_files=False, help="Importer votre taxonomie au format TTL. Vous pouvez utiliser les outils suivants pour les transformer...")  
    st.session_state["uploaded_file"] = uploaded_file  

    # Handle upload  
    if st.session_state["uploaded_file"] and not st.session_state["upload_bool"]:  
        st.session_state["upload_bool"] = True  
        print("importing")  
        ttl_data = import_ttl(uploaded_file)  
        st.session_state["ttl_data"] = ttl_data  

    # Dropdown for taxonomy selection   
    taxonomies = ["All", "NACE", "NACE-BEL", "EuroVoc", "STW", "Unesco", "Taxonomie du tourism", "Administrative territorial unit"]  
    selected_taxonomy = st.selectbox("Sélectionnez une taxonomie à mapper", taxonomies)  
    st.session_state["selected_taxonomy"] = selected_taxonomy  


    if chosen_function == "Alignement de taxonomies":  
        # Export mapping button  
        if st.button("Exporter le mapping"):  
            response_df = map_taxonomies(st.session_state["ttl_data"][1], st.session_state["selected_taxonomy"])
            print(response_df)  
            if response_df is not None:  
                # Save the DataFrame to an Excel file  
                excel_file = "mapping_result.xlsx"  
                #Add Excel template function
                #response_df.to_excel(excel_file, index=False)
                write_excel(response_df[["URI", "rdf:type", "align:entity1", "align:entity2", "align:relation", "align:measure^^xsd:float", "owl:annotatedProperty"]], response_df[["URI ", "skos:prefLabel", "skos:definition"]], response_df[["URI  ", "skos:prefLabel ", "skos:definition "]], excel_file)  
                # Provide a download button  
                with open(excel_file, "rb") as file:  
                    st.download_button(label="Télécharger le mapping", data=file, file_name=excel_file)  
        
    # Display existing messages in the Streamlit interface  
    for msg in st.session_state.history:  
        print(msg.role)
        if str(msg.role) == "AuthorRole.USER":  
            st.chat_message("human").write(msg.content)  
        elif str(msg.role) == "AuthorRole.ASSISTANT":  
            st.chat_message("ai").write(msg.content)
    
    # Set up markdown to fix chat_input and layer UI issues  
    st.markdown("""    
    <style>    
        .stChatInput {     
            position: fixed;    
            bottom: 50px;     
            width: 65%;    
            z-index: 3;    
            background-color: transparent; /* Make the input box background transparent */  
        }    
        .fixed-square {    
            position: fixed;    
            bottom: 0;    
            left: 28%;    
            width: 67%;    
            height: 100px;    
            background-color: transparent; /* Make the fixed square background transparent */  
            z-index: 2;    
        }    
        main {    
            z-index: 1;    
        }    
    </style>    
    <div class="fixed-square"></div>    
    """, unsafe_allow_html=True)


def map_taxonomies(ttl_data, selected_taxonomy):
    """
    Maps taxonomies based on the provided TTL data and selected taxonomy.  
  
    :param ttl_data: The TTL data containing taxonomy information.  
    :param selected_taxonomy: The selected taxonomy to map.  
    :return: DataFrame containing the mapped taxonomies.
    """
    mappings = []
    j = 1

    for index in ttl_data.index:
        if selected_taxonomy != "All":
            response, j = mapping_taxonomies(user_label=ttl_data["prefLabel"][index], user_definition=ttl_data["definition"][index], user_uri=ttl_data["subject"][index], j=j, taxonomy_filter=selected_taxonomy, loop=True)
        else:
            response, j = mapping_taxonomies(user_label=ttl_data["prefLabel"][index], user_definition=ttl_data["definition"][index], user_uri=ttl_data["subject"][index], j=j, taxonomy_filter=None, loop=True)
        
        mappings.extend(response)
    
    responses = pd.DataFrame.from_records(mappings)

    return responses

async def chat_and_map_taxonomies(user_input, ttl_data, selected_taxonomy, history):
    """
    Chats and maps taxonomies based on user input, TTL data, and selected taxonomy.  
  
    :param user_input: The user input for the chat.  
    :param ttl_data: The TTL data containing taxonomy information.  
    :param selected_taxonomy: The selected taxonomy to map.  
    :param history: The chat history.  
    :return: The response from the chat and mapping process.
    """
    #user_concept = ttl_data[(ttl_data['prefLabel'] == user_input) | (ttl_data['altLabel'] == user_input)]
    if selected_taxonomy != "All":
        response = chatting_and_mapping_taxonomies(user_input, ttl_data[0], selected_taxonomy, history)
    else:
        response = chatting_and_mapping_taxonomies(user_input, ttl_data[0], taxonomy_filter=None, history=history)

    return response


from typing import List, Optional
from langchain_core.messages import HumanMessage, AIMessage


async def chat_with_library(user_input: str) -> Optional[str]:
    """
    Chats with the library using the provided user input.

    :param user_input: The user input for the chat.
    :return: The response from the chat function, or None if no response is generated.
    """
    try:
        # Ensure chat history is initialized
        if "chat_history" not in st.session_state:
            st.session_state.chat_history = []

        # Append user query to chat history
        st.session_state.chat_history.append(HumanMessage(content=user_input))
        print("Updated Chat History:", st.session_state.chat_history)

        # Initialize the graph
        research_graph = build_and_test_graph()

        # Prepare the input state for the graph
        initial_state = {
            "messages": [
                ("user" if msg.type == "human" else "assistant", msg.content)
                for msg in st.session_state.chat_history
            ]
        }
        print(initial_state)

        # Process user input through the graph
        for result in research_graph.stream(
            initial_state, {"recursion_limit": 100}
        ):
            print("Intermediate State:", result)
            print("---")
            if "supervisor" in result.keys():
                if result["supervisor"]["next"] == END:
                    final_message = result["supervisor"]["messages"][-1].content
                    print("Final Response:", final_message)
                    return final_message
            elif "summarise" in result.keys():
                final_message = result["summarise"]["messages"][-1].content
                print("Final Response:", final_message)
                return final_message
            elif "chat" in result.keys():
                final_message = result["chat"]["messages"][-1].content
                print("Final Response:", final_message)
                return final_message

    except Exception as e:
        print("Error during chat_with_library execution:", str(e))
        return None



async def run_chatbot(kernel, user_input, chosen_function):
    
    st.chat_message("human").write(user_input)
    st.session_state["history"].add_user_message(user_input)
    
    try:
        if chosen_function == "Chat avec ma taxonomie":
            response = await chat_with_library(user_input)
        elif chosen_function == "Alignement de taxonomies":
            response = await chat_and_map_taxonomies(user_input, st.session_state["ttl_data"], st.session_state["selected_taxonomy"], st.session_state["history"])
        else:
            print(f"Intent chooser failed: {chosen_function}")
    except KeyError:
        response = "Please upload a model to use this function."

    st.chat_message("ai").write(response)
    print(st.session_state["history"])
    print(response)
    st.session_state["history"].add_assistant_message(str(response))
    st.session_state.chat_history.append(AIMessage(str(response)))



